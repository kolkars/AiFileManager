from pathlib import Path

from openpyxl import load_workbook

from .base import ContentUnit, ExtractionResult, normalize_text


class XlsxExtractor:
    def extract(self, path: Path) -> ExtractionResult:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            lines: list[str] = []
            units: list[ContentUnit] = []
            for sheet in workbook.worksheets:
                lines.append(sheet.title)
                units.append(ContentUnit("sheet", len(units), sheet.title, f"sheet:{sheet.title}", {"sheet": sheet.title}))
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    value = "\t".join("" if cell is None else str(cell) for cell in row)
                    lines.append(value)
                    units.append(ContentUnit("row", len(units), value, f"sheet:{sheet.title}/row:{row_number}", {"sheet": sheet.title, "row": row_number}))
            return ExtractionResult(normalize_text("\n".join(lines)), tuple(units), {"format": "xlsx", "sheet_count": len(workbook.sheetnames)})
        finally:
            workbook.close()
